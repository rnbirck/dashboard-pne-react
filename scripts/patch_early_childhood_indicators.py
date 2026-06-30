from __future__ import annotations

import argparse
import json
import math
import os
import re
import unicodedata
from collections import defaultdict
from pathlib import Path
from typing import Any

import psycopg2
from psycopg2.extras import execute_values
from openpyxl import load_workbook


REPO_DIR = Path(__file__).resolve().parents[1]
DEFAULT_DATA_DIR = REPO_DIR / "public" / "data"
DEFAULT_PYTHON_PROJECT_DIR = REPO_DIR / "data_pipeline"
DEFAULT_SINOPSE_DIR = (
    REPO_DIR.parent / "SENAI" / "DB" / "data" / "sinopse_estatistica_censo"
)

YEARS = range(2014, 2026)
UF_TARGET = "Rio Grande do Sul"

INDICATORS = {
    "creche": {
        "label": "Popula\u00e7\u00e3o de 0 a 3 anos que frequenta creche",
        "old_label": "Popula\u00e7\u00e3o de 0 a 3 anos que frequenta a escola/creche",
        "desc": "Percentual da popula\u00e7\u00e3o de 0 a 3 anos que frequenta creche.",
        "old_desc": "Percentual da popula\u00e7\u00e3o de 0 a 3 anos que frequenta a escola ou creche.",
        "numerator": "mat_creche_0_3",
        "denominator": "pop_0_3",
        "age_column": 5,
        "meta": {
            "pne_2014_2024": 50.0,
            "pne_2026_2036": 60.0,
        },
    },
    "pre_escola": {
        "label": "Popula\u00e7\u00e3o de 4 a 5 anos que frequenta a pr\u00e9-escola",
        "old_label": "Popula\u00e7\u00e3o de 4 a 5 anos que frequenta a escola/creche",
        "desc": "Percentual da popula\u00e7\u00e3o de 4 a 5 anos que frequenta a pr\u00e9-escola.",
        "old_desc": "Percentual da popula\u00e7\u00e3o de 4 a 5 anos que frequenta a escola ou creche.",
        "numerator": "mat_pre_escola_4_5",
        "denominator": "pop_4_5",
        "age_column": 6,
        "meta": {
            "pne_2014_2024": 100.0,
            "pne_2026_2036": 100.0,
        },
    },
}

CYCLES = {
    "pne_2014_2024": {
        "start_year": 2014,
        "end_year": 2024,
        "meta_label": "Meta PNE 2024",
    },
    "pne_2026_2036": {
        "start_year": 2015,
        "end_year": 2025,
        "meta_label": "Meta PNE 2036",
    },
}

DEMOGRAPHIC_STAGES = {
    "creche": ("infantil", "creche"),
    "pre_escola": ("infantil", "pre_escola"),
    "fundamental_anos_iniciais": (
        "fundamental_anos_iniciais",
        "fundamental_anos_iniciais",
    ),
    "fundamental_anos_finais": (
        "fundamental_anos_finais",
        "fundamental_anos_finais",
    ),
    "medio": ("medio", "medio"),
    "profissional": ("profissional", "educacao_profissional"),
    "eja": ("eja", "eja"),
}

RACE_ORDER = {
    "nao_declarada": "Não Declarada",
    "branca": "Branca",
    "preta": "Preta",
    "parda": "Parda",
    "amarela": "Amarela",
    "indigena": "Indígena",
}


def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(
        description=(
            "Atualiza os indicadores de creche e pre-escola do React usando "
            "as planilhas da Sinopse Estatistica do Censo Escolar."
        )
    )
    parser.add_argument("--data-dir", type=Path, default=DEFAULT_DATA_DIR)
    parser.add_argument(
        "--python-project-dir", type=Path, default=DEFAULT_PYTHON_PROJECT_DIR
    )
    parser.add_argument("--sinopse-dir", type=Path, default=DEFAULT_SINOPSE_DIR)
    return parser.parse_args()


def load_env_file(path: Path) -> None:
    if not path.exists():
        return
    for raw_line in path.read_text(encoding="utf-8").splitlines():
        line = raw_line.strip()
        if not line or line.startswith("#") or "=" not in line:
            continue
        key, value = line.split("=", 1)
        os.environ.setdefault(key.strip(), value.strip().strip('"').strip("'"))


def db_connection(python_project_dir: Path):
    load_env_file(python_project_dir / ".env")
    load_env_file(python_project_dir / "queries" / ".env")
    required = ["DB_USUARIO", "DB_SENHA", "DB_HOST", "DB_BANCO"]
    missing = [name for name in required if not os.getenv(name)]
    if missing:
        raise RuntimeError(
            "Variaveis de ambiente ausentes para Postgres: " + ", ".join(missing)
        )
    return psycopg2.connect(
        user=os.environ["DB_USUARIO"],
        password=os.environ["DB_SENHA"],
        host=os.environ["DB_HOST"],
        port=os.getenv("DB_PORT", "5432"),
        dbname=os.environ["DB_BANCO"],
    )


def locate_sinopse_file(sinopse_dir: Path, year: int) -> Path:
    candidates = sorted(sinopse_dir.glob(f"*{year}*.xlsx"))
    if not candidates:
        raise FileNotFoundError(f"Arquivo da sinopse nao encontrado para {year}")
    return candidates[0]


def sheet_for(year: int, indicator_key: str) -> str:
    if indicator_key == "creche":
        return "1.13" if year >= 2025 else "1.8"
    if indicator_key == "pre_escola":
        return "1.18" if year >= 2025 else "1.12"
    raise KeyError(indicator_key)


def number_or_none(value: Any) -> float | None:
    if value is None:
        return None
    try:
        number = float(value)
    except (TypeError, ValueError):
        return None
    if not math.isfinite(number):
        return None
    return number


def municipal_id(value: Any) -> str | None:
    number = number_or_none(value)
    if number is None:
        return None
    return str(int(number))


def normalize_label(value: Any) -> str:
    text = "" if value is None else str(value).strip()
    text = unicodedata.normalize("NFKD", text)
    text = "".join(char for char in text if not unicodedata.combining(char))
    text = re.sub(r"\s+", " ", text.lower()).strip()
    return text


def safe_cell_text(value: Any) -> str:
    return "" if value is None else str(value).strip()


def worksheet_title_text(worksheet) -> str:
    parts: list[str] = []
    for row in worksheet.iter_rows(min_row=1, max_row=4, values_only=True):
        parts.extend(safe_cell_text(value) for value in row if safe_cell_text(value))
    return " ".join(parts)


def demographic_stage_from_title(title: str) -> tuple[str, str] | None:
    normalized = normalize_label(title)
    if "educacao especial" in normalized or "educacao basica" in normalized:
        return None
    if "creche" in normalized:
        return DEMOGRAPHIC_STAGES["creche"]
    if "pre-escola" in normalized or "pre escola" in normalized:
        return DEMOGRAPHIC_STAGES["pre_escola"]
    if "anos iniciais" in normalized:
        return DEMOGRAPHIC_STAGES["fundamental_anos_iniciais"]
    if "anos finais" in normalized:
        return DEMOGRAPHIC_STAGES["fundamental_anos_finais"]
    if "ensino medio" in normalized:
        return DEMOGRAPHIC_STAGES["medio"]
    if "educacao profissional" in normalized:
        return DEMOGRAPHIC_STAGES["profissional"]
    if "jovens e adultos" in normalized or "eja" in normalized:
        return DEMOGRAPHIC_STAGES["eja"]
    return None


def race_key(value: Any) -> str | None:
    normalized = normalize_label(value)
    if not normalized or normalized == "total":
        return None
    normalized = normalized.replace("nao declarada", "nao_declarada")
    normalized = normalized.replace("indigena", "indigena")
    if normalized in RACE_ORDER:
        return normalized
    return None


def sex_key(value: Any) -> str | None:
    normalized = normalize_label(value)
    if normalized.startswith("feminino"):
        return "feminino"
    if normalized.startswith("masculino"):
        return "masculino"
    return None


def is_age_header(value: Any) -> bool:
    normalized = normalize_label(value)
    if not normalized:
        return False
    if normalized.startswith("faixa etaria"):
        return False
    if normalized.startswith("total"):
        return False
    return "anos" in normalized or normalized.startswith("ate ")


def sinopse_data_rows(worksheet):
    for row in worksheet.iter_rows(values_only=True):
        if len(row) < 5:
            continue
        if safe_cell_text(row[1]) != UF_TARGET:
            continue
        id_municipio = municipal_id(row[3])
        municipio = safe_cell_text(row[2])
        if id_municipio is None or not municipio:
            continue
        yield row, id_municipio, municipio


def age_columns(worksheet) -> list[tuple[int, str]]:
    best: list[tuple[int, str]] = []
    for row in worksheet.iter_rows(min_row=1, max_row=12, values_only=True):
        columns = [
            (index, safe_cell_text(value))
            for index, value in enumerate(row)
            if index >= 5 and is_age_header(value)
        ]
        if len(columns) > len(best):
            best = columns
    return best


def race_columns(worksheet) -> list[tuple[int, str, str]]:
    header_rows = list(worksheet.iter_rows(min_row=1, max_row=12, values_only=True))
    best_index = None
    best_count = 0
    for index, row in enumerate(header_rows):
        count = sum(1 for value in row if race_key(value) is not None)
        if count > best_count:
            best_count = count
            best_index = index
    if best_index is None or best_count == 0:
        return []

    header = header_rows[best_index]
    sex_header = header_rows[best_index - 1] if best_index > 0 else []
    current_sex = None
    sex_by_col: dict[int, str | None] = {}
    for col_index in range(len(header)):
        current = sex_key(sex_header[col_index] if col_index < len(sex_header) else None)
        if current:
            current_sex = current
        sex_by_col[col_index] = current_sex

    columns: list[tuple[int, str, str]] = []
    for col_index, value in enumerate(header):
        key = race_key(value)
        if key is None:
            continue
        sex = sex_by_col.get(col_index)
        if sex is None:
            continue
        columns.append((col_index, sex, RACE_ORDER[key]))
    return columns


def read_sinopse_demographic_rows(
    sinopse_dir: Path,
) -> tuple[list[dict[str, Any]], list[dict[str, Any]]]:
    faixa_rows: list[dict[str, Any]] = []
    cor_raca_rows: list[dict[str, Any]] = []

    for year in YEARS:
        workbook_path = locate_sinopse_file(sinopse_dir, year)
        workbook = load_workbook(workbook_path, read_only=True, data_only=True)
        try:
            for worksheet in workbook.worksheets:
                title = worksheet_title_text(worksheet)
                normalized_title = normalize_label(title)
                if "matriculas" not in normalized_title:
                    continue
                stage = demographic_stage_from_title(title)
                if stage is None:
                    continue
                etapa_ensino, secao_sinopse = stage
                fonte = (
                    "INEP - Sinopse Estatistica do Censo Escolar "
                    f"{year}, tabela {worksheet.title}"
                )

                if "faixa etaria" in normalized_title:
                    columns = age_columns(worksheet)
                    for row, id_municipio, municipio in sinopse_data_rows(worksheet):
                        for col_index, faixa_etaria in columns:
                            value = number_or_none(
                                row[col_index] if col_index < len(row) else None
                            )
                            faixa_rows.append(
                                {
                                    "ano": year,
                                    "id_municipio": id_municipio,
                                    "municipio": municipio,
                                    "etapa_ensino": etapa_ensino,
                                    "secao_sinopse": secao_sinopse,
                                    "faixa_etaria": faixa_etaria.strip(),
                                    "matriculas": int(value or 0),
                                    "fonte": fonte,
                                }
                            )
                elif "sexo e cor/raca" in normalized_title:
                    columns = race_columns(worksheet)
                    for row, id_municipio, municipio in sinopse_data_rows(worksheet):
                        for col_index, sexo, cor_raca in columns:
                            value = number_or_none(
                                row[col_index] if col_index < len(row) else None
                            )
                            cor_raca_rows.append(
                                {
                                    "ano": year,
                                    "id_municipio": id_municipio,
                                    "municipio": municipio,
                                    "etapa_ensino": etapa_ensino,
                                    "secao_sinopse": secao_sinopse,
                                    "sexo": sexo,
                                    "cor_raca": cor_raca,
                                    "matriculas": int(value or 0),
                                    "fonte": fonte,
                                }
                            )
        finally:
            workbook.close()

    return faixa_rows, cor_raca_rows


def sync_sinopse_demographic_tables(
    python_project_dir: Path,
    faixa_rows: list[dict[str, Any]],
    cor_raca_rows: list[dict[str, Any]],
) -> None:
    with db_connection(python_project_dir) as connection:
        with connection.cursor() as cursor:
            cursor.execute(
                """
                CREATE TABLE IF NOT EXISTS matriculas_faixa_etaria (
                    ano bigint,
                    id_municipio text,
                    municipio text,
                    etapa_ensino text,
                    secao_sinopse text,
                    faixa_etaria text,
                    matriculas bigint,
                    fonte text,
                    data_carga timestamp without time zone DEFAULT now()
                )
                """
            )
            cursor.execute(
                """
                CREATE TABLE IF NOT EXISTS matriculas_cor_raca (
                    ano bigint,
                    id_municipio text,
                    municipio text,
                    etapa_ensino text,
                    secao_sinopse text,
                    sexo text,
                    cor_raca text,
                    matriculas bigint,
                    fonte text,
                    data_carga timestamp without time zone DEFAULT now()
                )
                """
            )
            cursor.execute("DELETE FROM matriculas_faixa_etaria WHERE ano BETWEEN %s AND %s", (min(YEARS), max(YEARS)))
            cursor.execute("DELETE FROM matriculas_cor_raca WHERE ano BETWEEN %s AND %s", (min(YEARS), max(YEARS)))

            if faixa_rows:
                execute_values(
                    cursor,
                    """
                    INSERT INTO matriculas_faixa_etaria (
                        ano, id_municipio, municipio, etapa_ensino, secao_sinopse,
                        faixa_etaria, matriculas, fonte, data_carga
                    ) VALUES %s
                    """,
                    [
                        (
                            row["ano"],
                            row["id_municipio"],
                            row["municipio"],
                            row["etapa_ensino"],
                            row["secao_sinopse"],
                            row["faixa_etaria"],
                            row["matriculas"],
                            row["fonte"],
                        )
                        for row in faixa_rows
                    ],
                    template="(%s, %s, %s, %s, %s, %s, %s, %s, now())",
                    page_size=5000,
                )
            if cor_raca_rows:
                execute_values(
                    cursor,
                    """
                    INSERT INTO matriculas_cor_raca (
                        ano, id_municipio, municipio, etapa_ensino, secao_sinopse,
                        sexo, cor_raca, matriculas, fonte, data_carga
                    ) VALUES %s
                    """,
                    [
                        (
                            row["ano"],
                            row["id_municipio"],
                            row["municipio"],
                            row["etapa_ensino"],
                            row["secao_sinopse"],
                            row["sexo"],
                            row["cor_raca"],
                            row["matriculas"],
                            row["fonte"],
                        )
                        for row in cor_raca_rows
                    ],
                    template="(%s, %s, %s, %s, %s, %s, %s, %s, %s, now())",
                    page_size=5000,
                )

            cursor.execute(
                """
                CREATE OR REPLACE VIEW vw_educacao_matriculas_faixa_etaria AS
                SELECT
                    ano::integer AS ano,
                    id_municipio::text AS id_municipio,
                    municipio::text AS municipio,
                    etapa_ensino::text AS etapa_ensino,
                    secao_sinopse::text AS secao_sinopse,
                    faixa_etaria::text AS faixa_etaria,
                    matriculas::bigint AS matriculas,
                    fonte::text AS fonte,
                    data_carga
                FROM matriculas_faixa_etaria
                """
            )
            cursor.execute(
                """
                CREATE OR REPLACE VIEW vw_educacao_matriculas_cor_raca AS
                SELECT
                    ano::integer AS ano,
                    id_municipio::text AS id_municipio,
                    municipio::text AS municipio,
                    etapa_ensino::text AS etapa_ensino,
                    secao_sinopse::text AS secao_sinopse,
                    sexo::text AS sexo,
                    cor_raca::text AS cor_raca,
                    matriculas::bigint AS matriculas,
                    fonte::text AS fonte,
                    data_carga
                FROM matriculas_cor_raca
                """
            )
            cursor.execute(
                """
                CREATE INDEX IF NOT EXISTS idx_matriculas_faixa_etaria_lookup
                ON matriculas_faixa_etaria (id_municipio, ano, etapa_ensino, secao_sinopse)
                """
            )
            cursor.execute(
                """
                CREATE INDEX IF NOT EXISTS idx_matriculas_cor_raca_lookup
                ON matriculas_cor_raca (id_municipio, ano, etapa_ensino, secao_sinopse, cor_raca)
                """
            )
        connection.commit()


def read_sinopse_enrollments(
    sinopse_dir: Path,
) -> dict[tuple[int, str], dict[str, float]]:
    enrollments: dict[tuple[int, str], dict[str, float]] = defaultdict(dict)

    for year in YEARS:
        workbook_path = locate_sinopse_file(sinopse_dir, year)
        workbook = load_workbook(workbook_path, read_only=True, data_only=True)
        try:
            for indicator_key, config in INDICATORS.items():
                worksheet = workbook[sheet_for(year, indicator_key)]
                target_col = int(config["age_column"])
                numerator = str(config["numerator"])

                for row in worksheet.iter_rows(values_only=True):
                    if len(row) <= target_col:
                        continue
                    if str(row[1]).strip() != UF_TARGET:
                        continue
                    id_municipio = municipal_id(row[3])
                    if id_municipio is None:
                        continue
                    value = number_or_none(row[target_col])
                    enrollments[(year, id_municipio)][numerator] = value or 0.0
        finally:
            workbook.close()

    return enrollments


def read_population(
    python_project_dir: Path,
) -> dict[tuple[int, str], dict[str, Any]]:
    query = """
        SELECT
            p.ano,
            m.id_municipio::text AS id_municipio,
            m.municipio,
            SUM(CASE WHEN p.idade <= 3 THEN p.pop_estimada ELSE 0 END) AS pop_0_3,
            SUM(CASE WHEN p.idade BETWEEN 4 AND 5 THEN p.pop_estimada ELSE 0 END) AS pop_4_5
        FROM populacao_idade_rs p
        JOIN municipios m ON p.id_municipio = m.id_municipio
        WHERE p.ano BETWEEN 2014 AND 2025
        GROUP BY p.ano, m.id_municipio, m.municipio
        ORDER BY p.ano, m.municipio
    """
    rows: dict[tuple[int, str], dict[str, Any]] = {}
    with db_connection(python_project_dir) as connection:
        with connection.cursor() as cursor:
            cursor.execute(query)
            for year, id_municipio, municipio, pop_0_3, pop_4_5 in cursor.fetchall():
                rows[(int(year), str(id_municipio))] = {
                    "municipio": municipio,
                    "pop_0_3": float(pop_0_3 or 0),
                    "pop_4_5": float(pop_4_5 or 0),
                }
    return rows


def build_rates(
    enrollments: dict[tuple[int, str], dict[str, float]],
    population: dict[tuple[int, str], dict[str, Any]],
) -> dict[str, dict[str, list[dict[str, float | int]]]]:
    rates: dict[str, dict[str, list[dict[str, float | int]]]] = defaultdict(
        lambda: defaultdict(list)
    )
    for key, pop_row in population.items():
        year, _id_municipio = key
        enrollment_row = enrollments.get(key, {})
        municipio = str(pop_row["municipio"])
        for indicator_key, config in INDICATORS.items():
            numerator = float(enrollment_row.get(str(config["numerator"]), 0.0))
            denominator = float(pop_row.get(str(config["denominator"]), 0.0))
            value = (numerator / denominator * 100.0) if denominator else 0.0
            rates[municipio][indicator_key].append({"ano": year, "valor": value})

    for municipio_rates in rates.values():
        for series in municipio_rates.values():
            series.sort(key=lambda row: int(row["ano"]))
    return rates


def fmt_number(value: float | int | None, decimals: int = 1, signed: bool = False) -> str:
    if value is None:
        return "-"
    number = float(value)
    prefix = "+" if signed and number > 0 else ""
    formatted = f"{number:.{decimals}f}"
    return prefix + formatted.replace(".", ",")


def fmt_percent(value: float | int | None) -> str:
    return f"{fmt_number(value)}%"


def fmt_signed_pp(value: float | int | None) -> str:
    if value is None:
        return "-"
    return f"{fmt_number(value, signed=True)} p.p."


def variation_text(delta: float | None) -> str:
    if delta is None:
        return "-"
    if delta > 0:
        return f"Alta de {fmt_number(delta)} p.p."
    if delta < 0:
        return f"Queda de {fmt_number(abs(delta))} p.p."
    return "Sem varia\u00e7\u00e3o"


def build_display(result: dict[str, Any], label_item: dict[str, str]) -> dict[str, str]:
    if not result.get("available"):
        return {
            "start_value": "-",
            "end_value": "-",
            "variation": "-",
            "distance": "-",
            "status": "Sem dados",
            "interpretation": (
                "N\u00e3o foi poss\u00edvel montar a compara\u00e7\u00e3o desse indicador "
                "com os dados dispon\u00edveis para o munic\u00edpio selecionado."
            ),
        }

    end_value = fmt_percent(result["end_value"])
    meta_value = fmt_percent(result["meta"])
    distance = fmt_signed_pp(result["distance"])
    status = "Meta atingida" if result["atingida"] else "Meta n\u00e3o atingida"

    if result["atingida"]:
        interpretation = (
            f"Em {result['end_year']}, o munic\u00edpio alcan\u00e7ou {end_value} "
            f"e superou a meta definida ({meta_value}). O desempenho ficou "
            f"{distance} acima da refer\u00eancia."
        )
    else:
        interpretation = (
            f"Em {result['end_year']}, o munic\u00edpio chegou a {end_value}, "
            f"mas n\u00e3o alcan\u00e7ou a meta definida ({meta_value}). A dist\u00e2ncia "
            f"atual para a refer\u00eancia \u00e9 de {distance}."
        )

    return {
        "start_value": fmt_percent(result["start_value"]),
        "end_value": end_value,
        "variation": variation_text(result["raw_delta"]),
        "distance": distance,
        "status": status,
        "interpretation": interpretation,
    }


def select_reference_rows(
    series: list[dict[str, float | int]],
    target_start_year: int,
    target_end_year: int,
) -> tuple[dict[str, float | int] | None, dict[str, float | int] | None]:
    if not series:
        return None, None
    start_candidates = [row for row in series if int(row["ano"]) == target_start_year]
    if not start_candidates:
        start_candidates = [row for row in series if int(row["ano"]) >= target_start_year]
    start_row = start_candidates[0] if start_candidates else series[0]

    end_candidates = [row for row in series if int(row["ano"]) == target_end_year]
    if not end_candidates:
        end_candidates = [row for row in series if int(row["ano"]) <= target_end_year]
    end_row = end_candidates[-1] if end_candidates else series[-1]

    if int(start_row["ano"]) > int(end_row["ano"]):
        start_row, end_row = series[0], series[-1]
    if len(series) > 1 and int(start_row["ano"]) == int(end_row["ano"]):
        start_row, end_row = series[0], series[-1]
    return start_row, end_row


def build_result(
    source_series: list[dict[str, float | int]],
    indicator_key: str,
    cycle_key: str,
) -> dict[str, Any]:
    cycle = CYCLES[cycle_key]
    meta = INDICATORS[indicator_key]["meta"][cycle_key]
    series = [
        {"ano": int(row["ano"]), "valor": float(row["valor"])}
        for row in source_series
        if cycle["start_year"] <= int(row["ano"]) <= cycle["end_year"]
    ]
    start_row, end_row = select_reference_rows(
        series, int(cycle["start_year"]), int(cycle["end_year"])
    )
    if start_row is None or end_row is None:
        return {
            "available": False,
            "meta": meta,
            "meta_label": cycle["meta_label"],
            "direction": "at_least",
            "start_year": None,
            "end_year": None,
            "start_value": None,
            "end_value": None,
            "raw_delta": None,
            "progress_delta": None,
            "distance": None,
            "atingida": False,
            "tracks_goal": True,
            "display": build_display({"available": False}, INDICATORS[indicator_key]),
        }

    start_value = float(start_row["valor"])
    end_value = float(end_row["valor"])
    distance = end_value - meta
    result = {
        "available": True,
        "start_year": int(start_row["ano"]),
        "end_year": int(end_row["ano"]),
        "start_value": start_value,
        "end_value": end_value,
        "raw_delta": end_value - start_value,
        "progress_delta": end_value - start_value,
        "meta": meta,
        "meta_label": cycle["meta_label"],
        "direction": "at_least",
        "distance": distance,
        "atingida": distance >= 0,
        "tracks_goal": True,
        "series": [
            row
            for row in series
            if int(start_row["ano"]) <= int(row["ano"]) <= int(end_row["ano"])
        ],
    }
    result["display"] = build_display(result, INDICATORS[indicator_key])
    return result


def replace_known_text(value: Any) -> Any:
    if isinstance(value, str):
        for config in INDICATORS.values():
            value = value.replace(str(config["old_label"]), str(config["label"]))
            value = value.replace(str(config["old_desc"]), str(config["desc"]))
        return value
    if isinstance(value, list):
        return [replace_known_text(item) for item in value]
    if isinstance(value, dict):
        return {key: replace_known_text(child) for key, child in value.items()}
    return value


def update_indicator_metadata(payload: dict[str, Any]) -> None:
    for cycle in payload.get("cycles", {}).values():
        for category in cycle.get("categories", []):
            for item in category.get("items", []):
                key = item.get("key")
                if key in INDICATORS:
                    item["label"] = INDICATORS[key]["label"]
                    item["desc"] = INDICATORS[key]["desc"]


def ranking_row(category_key: str, indicator_key: str, result: dict[str, Any]) -> dict[str, Any]:
    config = INDICATORS.get(indicator_key, {})
    return {
        "indicator_key": indicator_key,
        "label": config.get("label", indicator_key),
        "sub": "",
        "category": category_key,
        "progress_delta": result.get("progress_delta"),
        "distance": result.get("distance"),
        "atingida": result.get("atingida"),
        "available": result.get("available"),
        "value_mode": "percent",
        "start_year": result.get("start_year"),
        "end_year": result.get("end_year"),
        "start_value": result.get("start_value"),
        "end_value": result.get("end_value"),
        "display": result.get("display", {}),
    }


def update_ranking_rows(rows: list[dict[str, Any]], updates: dict[str, dict[str, Any]]) -> None:
    for index, row in enumerate(rows):
        indicator_key = row.get("indicator_key")
        if indicator_key in updates:
            updated = dict(row)
            result = updates[indicator_key]
            updated.update(
                {
                    "label": INDICATORS[indicator_key]["label"],
                    "progress_delta": result.get("progress_delta"),
                    "distance": result.get("distance"),
                    "atingida": result.get("atingida"),
                    "available": result.get("available"),
                    "start_year": result.get("start_year"),
                    "end_year": result.get("end_year"),
                    "start_value": result.get("start_value"),
                    "end_value": result.get("end_value"),
                    "display": result.get("display", {}),
                }
            )
            rows[index] = updated


def recompute_atendimento_rankings(cycle_payload: dict[str, Any]) -> None:
    rankings = cycle_payload.get("rankings", {})
    atendimento = rankings.get("atendimento")
    indicadores = cycle_payload.get("indicadores", {})
    if not isinstance(atendimento, dict) or not isinstance(indicadores, dict):
        return

    existing_by_key: dict[str, dict[str, Any]] = {}
    for list_name in ("top_avancos", "top_atencao"):
        for row in atendimento.get(list_name, []):
            if "indicator_key" in row:
                existing_by_key[row["indicator_key"]] = row

    rows = []
    for indicator_key, result in indicadores.items():
        if not result.get("available") or not result.get("tracks_goal", True):
            continue
        if result.get("start_year") == result.get("end_year"):
            continue
        row = dict(existing_by_key.get(indicator_key, {}))
        if not row:
            row = ranking_row("atendimento", indicator_key, result)
        row.update(
            {
                "progress_delta": result.get("progress_delta"),
                "distance": result.get("distance"),
                "atingida": result.get("atingida"),
                "available": result.get("available"),
                "start_year": result.get("start_year"),
                "end_year": result.get("end_year"),
                "start_value": result.get("start_value"),
                "end_value": result.get("end_value"),
                "display": result.get("display", {}),
            }
        )
        if indicator_key in INDICATORS:
            row["label"] = INDICATORS[indicator_key]["label"]
        rows.append(row)

    atendimento["top_avancos"] = sorted(
        rows,
        key=lambda row: float(row.get("progress_delta") or 0),
        reverse=True,
    )[:3]
    atendimento["top_atencao"] = sorted(
        [row for row in rows if not row.get("atingida")],
        key=lambda row: float(row.get("distance") or 0),
    )[:3]


def recompute_diagnostic(diagnostic: dict[str, Any]) -> None:
    categories = diagnostic.get("categories")
    if not isinstance(categories, list):
        return

    all_goal_records = []
    all_attention = []
    all_positive = []

    for category in categories:
        indicators = category.get("indicators", [])
        goal_records = [
            item for item in indicators if item.get("tracks_goal", True)
        ]
        positive = [
            item for item in goal_records if item.get("result", {}).get("atingida")
        ]
        attention = [
            item
            for item in goal_records
            if item.get("result", {}).get("available")
            and not item.get("result", {}).get("atingida")
        ]
        informative = [
            item for item in indicators if not item.get("tracks_goal", True)
        ]

        category["goal_total"] = len(goal_records)
        category["informative_total"] = len(informative)
        category["achieved"] = len(positive)
        category["attention_count"] = len(attention)
        category["attention_indicators"] = [item.get("key") for item in attention]
        category["positive_indicators"] = [item.get("key") for item in positive]
        category["informative_indicators"] = [item.get("key") for item in informative]
        category["counter_text"] = (
            f"{len(positive)} de {len(goal_records)} metas atingidas; "
            f"{len(informative)} informativos"
        )
        sorted_attention = sorted(
            attention,
            key=lambda item: float(item.get("result", {}).get("distance") or 0),
        )
        evidence_lines = []
        for item in sorted_attention[:3]:
            evidence_lines.append(
                f"{item.get('label')}: "
                f"{item.get('result', {}).get('display', {}).get('end_value', '-')}, "
                "abaixo da meta acompanhada."
            )
        evidence_lines.append(
            f"A dimens\u00e3o atingiu {len(positive)} de {len(goal_records)} metas acompanhadas."
        )
        category["evidence_lines"] = evidence_lines

        all_goal_records.extend(goal_records)
        all_attention.extend(attention)
        all_positive.extend(positive)

    diagnostic["summary"] = {
        "indicadores_analisados": len(all_goal_records),
        "metas_atingidas": len(all_positive),
        "pontos_de_atencao": len(all_attention),
    }
    diagnostic["principais_desafios"] = [
        f"Ampliar a cobertura em {item.get('label')}."
        for item in sorted(
            all_attention,
            key=lambda item: float(item.get("result", {}).get("distance") or 0),
        )[:3]
    ]
    diagnostic["pontos_positivos"] = [
        f"Meta alcan\u00e7ada: {item.get('label')}."
        for item in sorted(
            all_positive,
            key=lambda item: float(item.get("result", {}).get("distance") or 0),
            reverse=True,
        )[:3]
    ]


def patch_municipio_file(path: Path, rates: dict[str, dict[str, list[dict[str, float | int]]]]) -> bool:
    payload = json.loads(path.read_text(encoding="utf-8"))
    payload = replace_known_text(payload)
    municipio = payload.get("municipio")
    municipio_rates = rates.get(str(municipio), {})
    if not municipio_rates:
        return False

    changed = False
    for cycle_key in CYCLES:
        cycle_payload = payload.get(cycle_key, {})
        indicadores = cycle_payload.get("indicadores", {})
        updates = {}
        for indicator_key in INDICATORS:
            if indicator_key not in indicadores:
                continue
            result = build_result(
                municipio_rates.get(indicator_key, []),
                indicator_key,
                cycle_key,
            )
            indicadores[indicator_key] = result
            updates[indicator_key] = result
            changed = True

        rankings = cycle_payload.get("rankings", {})
        for category in rankings.values():
            if not isinstance(category, dict):
                continue
            update_ranking_rows(category.get("top_avancos", []), updates)
            update_ranking_rows(category.get("top_atencao", []), updates)
        recompute_atendimento_rankings(cycle_payload)

        diagnostic = cycle_payload.get("diagnostico", {})
        if isinstance(diagnostic, dict):
            for category in diagnostic.get("categories", []):
                for item in category.get("indicators", []):
                    key = item.get("key")
                    if key in INDICATORS:
                        item["label"] = INDICATORS[key]["label"]
                        item["desc"] = INDICATORS[key]["desc"]
                    if key in updates:
                        item["result"] = updates[key]
                        item["achieved"] = updates[key].get("atingida", False)
            recompute_diagnostic(diagnostic)

    if changed:
        path.write_text(
            json.dumps(payload, ensure_ascii=False, indent=2) + "\n",
            encoding="utf-8",
        )
    return changed


def patch_global_indicators(data_dir: Path) -> None:
    path = data_dir / "indicadores.json"
    if not path.exists():
        return
    payload = json.loads(path.read_text(encoding="utf-8"))
    payload = replace_known_text(payload)
    update_indicator_metadata(payload)
    path.write_text(
        json.dumps(payload, ensure_ascii=False, indent=2) + "\n",
        encoding="utf-8",
    )


def main() -> int:
    args = parse_args()
    data_dir = args.data_dir.resolve()
    sinopse_dir = args.sinopse_dir.resolve()
    python_project_dir = args.python_project_dir.resolve()

    if not data_dir.exists():
        raise FileNotFoundError(f"Diretorio de dados nao encontrado: {data_dir}")
    if not sinopse_dir.exists():
        raise FileNotFoundError(f"Diretorio de sinopses nao encontrado: {sinopse_dir}")

    print("[early-childhood] Lendo matriculas das sinopses...")
    enrollments = read_sinopse_enrollments(sinopse_dir)
    print(f"[early-childhood] Linhas ano/municipio com matriculas: {len(enrollments)}")

    print("[early-childhood] Lendo recortes de faixa etaria e cor/raca das sinopses...")
    faixa_rows, cor_raca_rows = read_sinopse_demographic_rows(sinopse_dir)
    print(
        "[early-childhood] Linhas de detalhe extraidas: "
        f"{len(faixa_rows)} faixa etaria; {len(cor_raca_rows)} cor/raca"
    )

    print("[early-childhood] Atualizando tabelas/views demograficas no Postgres local...")
    sync_sinopse_demographic_tables(python_project_dir, faixa_rows, cor_raca_rows)

    print("[early-childhood] Lendo populacao do Postgres local...")
    population = read_population(python_project_dir)
    print(f"[early-childhood] Linhas ano/municipio com populacao: {len(population)}")

    rates = build_rates(enrollments, population)
    patch_global_indicators(data_dir)

    municipios_dir = data_dir / "municipios"
    patched = 0
    for path in sorted(municipios_dir.glob("*/index.json")):
        if patch_municipio_file(path, rates):
            patched += 1

    print(f"[early-childhood] Municipios atualizados: {patched}")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
