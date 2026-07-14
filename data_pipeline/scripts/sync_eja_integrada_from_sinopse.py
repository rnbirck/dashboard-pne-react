"""Sincroniza o indicador de EJA a partir das tabelas oficiais da Sinopse.

Os campos são resolvidos pelo texto normalizado dos cabeçalhos multinível. A
posição física da coluna não faz parte do contrato do importador.
"""

from __future__ import annotations

import argparse
from collections import Counter
import hashlib
import json
import re
import sys
import tempfile
import unicodedata
import zipfile
from pathlib import Path

import pandas as pd
from openpyxl import load_workbook
from sqlalchemy import text

BASE_DIR = Path(__file__).resolve().parents[1]
if str(BASE_DIR) not in sys.path:
    sys.path.insert(0, str(BASE_DIR))

from src.data.repository import get_local_postgres_engine  # noqa: E402
from src.eja_integrada_indicator import (  # noqa: E402
    calculate_eja_integrada_series,
    validate_dependency_totals,
)

EXPECTED_RS_MUNICIPALITIES = 497
DEPENDENCIES = ("federal", "estadual", "municipal", "privada")


def normalize(value: object) -> str:
    text_value = unicodedata.normalize("NFKD", str(value or ""))
    text_value = "".join(char for char in text_value if not unicodedata.combining(char))
    text_value = re.sub(r"(?<=[A-Za-z])(?=\d)|(?<=\d)(?=[A-Za-z])", " ", text_value)
    return " ".join(re.sub(r"[^a-zA-Z0-9]+", " ", text_value).lower().split())


def sha256(path: Path) -> str:
    digest = hashlib.sha256()
    with path.open("rb") as source:
        for chunk in iter(lambda: source.read(1024 * 1024), b""):
            digest.update(chunk)
    return digest.hexdigest()


def _sheet_name(workbook, suffix: str) -> str:
    candidates = [name for name in workbook.sheetnames if normalize(name).endswith(normalize(suffix))]
    if len(candidates) != 1:
        raise ValueError(f"Esperada uma planilha {suffix}; encontradas {candidates}.")
    return candidates[0]


def _semantic_sheet_name(workbook, required_title_terms: tuple[str, ...]) -> str:
    required = tuple(normalize(term) for term in required_title_terms)
    candidates = []
    for name in workbook.sheetnames:
        sheet = workbook[name]
        title = " ".join(
            normalize(sheet.cell(row, 1).value) for row in range(1, 6)
        )
        if all(f" {term} " in f" {title} " for term in required):
            candidates.append(name)
    if len(candidates) != 1:
        raise ValueError(
            "Esperada uma única planilha pelos termos semânticos "
            f"{required_title_terms}; encontradas {candidates}."
        )
    return candidates[0]


def _extract_workbook(
    source: Path,
    suffixes: tuple[str, ...],
    temp_dir: Path,
    *,
    title_terms: tuple[str, ...] = (),
) -> Path:
    if source.suffix.lower() == ".xlsx":
        return source
    if source.suffix.lower() != ".zip":
        raise ValueError("A fonte deve ser o XLSX ou ZIP oficial da Sinopse.")
    with zipfile.ZipFile(source) as archive:
        for member in archive.namelist():
            if not member.lower().endswith(".xlsx"):
                continue
            target = temp_dir / Path(member).name
            with archive.open(member) as origin, target.open("wb") as output:
                output.write(origin.read())
            workbook = load_workbook(target, read_only=True, data_only=True)
            try:
                suffixes_match = all(
                    any(
                        normalize(name).endswith(normalize(s))
                        for name in workbook.sheetnames
                    )
                    for s in suffixes
                )
                title_matches = True
                if title_terms:
                    try:
                        _semantic_sheet_name(workbook, title_terms)
                    except ValueError:
                        title_matches = False
                if suffixes_match and title_matches:
                    return target
            finally:
                workbook.close()
            target.unlink(missing_ok=True)
    raise ValueError(f"Nenhum XLSX contém as planilhas {suffixes}.")


def _integer(value: object, field: str, row: int) -> int:
    if value is None or str(value).strip() == "":
        raise ValueError(f"Valor ausente em {field}, linha {row}.")
    numeric = int(float(value))
    if numeric < 0:
        raise ValueError(f"Valor negativo em {field}, linha {row}.")
    return numeric


def _discover_data_coordinates(sheet, uf: str) -> tuple[int, int, int, int]:
    target_uf = normalize(uf)
    coordinate_pairs: Counter[tuple[int, int]] = Counter()
    first_code_rows: dict[int, int] = {}
    first_country_row: int | None = None
    last_data_column = 0
    for row_number, row in enumerate(sheet.iter_rows(values_only=True), 1):
        if first_country_row is None and row and normalize(row[0]) == "brasil":
            first_country_row = row_number
        uf_columns = [
            index + 1 for index, value in enumerate(row) if normalize(value) == target_uf
        ]
        code_columns = [
            index + 1
            for index, value in enumerate(row)
            if re.fullmatch(r"\d{7}", str(value or "").strip())
        ]
        for code_column in code_columns:
            first_code_rows.setdefault(code_column, row_number)
        for uf_column in uf_columns:
            for code_column in code_columns:
                pair = (code_column, uf_column)
                coordinate_pairs[pair] += 1
                last_data_column = max(
                    last_data_column,
                    max(
                        (
                            index + 1
                            for index, value in enumerate(row)
                            if value is not None and str(value).strip()
                        ),
                        default=0,
                    ),
                )
    if not coordinate_pairs:
        raise ValueError("Não foi possível localizar semanticamente código IBGE e UF.")
    (code_column, uf_column), matches = coordinate_pairs.most_common(1)[0]
    if matches < 3:
        raise ValueError("Código IBGE e UF não aparecem juntos em linhas suficientes.")
    data_row = min(first_code_rows[code_column], first_country_row or sheet.max_row)
    return data_row, code_column, uf_column, last_data_column


def _header_paths(
    sheet, last_header_row: int, last_data_column: int | None = None
) -> dict[int, str]:
    # In read-only mode openpyxl exposes the merged region only in its
    # top-left cell.  Propagate header labels horizontally and vertically so
    # each physical column receives its complete semantic path without ever
    # depending on a fixed Excel letter/position.
    raw_rows = [
        list(row)
        for row in sheet.iter_rows(
            min_row=1,
            max_row=last_header_row,
            max_col=last_data_column or sheet.max_column,
            values_only=True,
        )
    ]
    structural_start = next(
        (
            index
            for index, row in enumerate(raw_rows)
            if any("codigo do municipio" in normalize(value) for value in row)
        ),
        0,
    )
    raw_rows = raw_rows[structural_start:]
    expanded_rows: list[list[object]] = []
    max_column = last_data_column or sheet.max_column
    for row_index, raw_row in enumerate(raw_rows):
        expanded: list[object] = []
        previous = None
        for column_index, value in enumerate(raw_row):
            ancestor_boundary = any(
                raw_rows[ancestor][column_index] is not None
                and str(raw_rows[ancestor][column_index]).strip()
                for ancestor in range(row_index)
            )
            if ancestor_boundary:
                previous = None
            if value is not None and str(value).strip():
                previous = value
            expanded.append(previous)
        expanded_rows.append(expanded)
    paths = {}
    for column in range(1, max_column + 1):
        parts = []
        for row in range(len(expanded_rows)):
            value = expanded_rows[row][column - 1]
            part = normalize(value)
            if part and (not parts or parts[-1] != part):
                parts.append(part)
        paths[column] = " | ".join(parts)
    return paths


def _column(paths: dict[int, str], *, required: tuple[str, ...], forbidden: tuple[str, ...] = ()) -> int:
    required_normalized = tuple(normalize(item) for item in required)
    forbidden_normalized = tuple(normalize(item) for item in forbidden)
    matches = [
        column
        for column, path in paths.items()
        if all(f" {term} " in f" {path} " for term in required_normalized)
        and not any(f" {term} " in f" {path} " for term in forbidden_normalized)
    ]
    if len(matches) != 1:
        evidence = {column: paths[column] for column in matches[:10]}
        raise ValueError(
            f"Cabeçalho ambíguo/ausente para {required}: colunas={evidence}."
        )
    return matches[0]


def _dependency_columns(paths, category: tuple[str, ...]) -> dict[str, int]:
    columns = {
        dependency: _column(paths, required=category + (dependency,))
        for dependency in ("federal", "estadual", "municipal")
    }
    columns["privada"] = _column(
        paths,
        required=category + ("privada",),
        forbidden=("categoria da escola privada",),
    )
    return columns


def _parse_sheet(
    workbook_path: Path,
    suffix: str | None,
    year: int,
    uf: str,
    field_specs: dict,
    *,
    title_terms: tuple[str, ...] = (),
) -> tuple[pd.DataFrame, dict]:
    workbook = load_workbook(workbook_path, read_only=True, data_only=True)
    try:
        name = (
            _semantic_sheet_name(workbook, title_terms)
            if title_terms
            else _sheet_name(workbook, str(suffix))
        )
        sheet = workbook[name]
        data_row, code_column, uf_column, last_data_column = (
            _discover_data_coordinates(sheet, uf)
        )
        paths = _header_paths(sheet, data_row - 1, last_data_column)
        resolved = {}
        for field, spec in field_specs.items():
            if spec.get("dependency_category"):
                resolved[field] = _dependency_columns(paths, spec["dependency_category"])
            elif spec.get("sum_columns"):
                resolved[field] = [
                    _column(paths, required=tuple(terms))
                    for terms in spec["sum_columns"]
                ]
            else:
                resolved[field] = _column(
                    paths,
                    required=spec["required"],
                    forbidden=spec.get("forbidden", ()),
                )
        records = []
        for row_number, values in enumerate(
            sheet.iter_rows(min_row=data_row, values_only=True), data_row
        ):
            if normalize(values[uf_column - 1]) != normalize(uf):
                continue
            code = str(values[code_column - 1] or "").strip()
            if not re.fullmatch(r"\d{7}", code):
                continue
            record = {"ano": year, "id_municipio": code}
            for field, column in resolved.items():
                if isinstance(column, dict):
                    for dependency, dependency_column in column.items():
                        record[f"{field}_{dependency}"] = _integer(
                            values[dependency_column - 1], f"{field}_{dependency}", row_number
                        )
                elif isinstance(column, list):
                    record[field] = sum(
                        _integer(values[item - 1], field, row_number) for item in column
                    )
                else:
                    record[field] = _integer(values[column - 1], field, row_number)
            records.append(record)
    finally:
        workbook.close()
    frame = pd.DataFrame(records)
    if len(frame) != EXPECTED_RS_MUNICIPALITIES:
        raise ValueError(f"Cobertura de {name}: {len(frame)}; esperado {EXPECTED_RS_MUNICIPALITIES}.")
    if frame.duplicated(["ano", "id_municipio"]).any():
        raise ValueError(f"Códigos IBGE duplicados na planilha {name}.")
    return frame, {"sheet": name, "resolved_columns": resolved}


def parse_sources(eja_source: Path, professional_source: Path, year: int, uf: str) -> tuple[pd.DataFrame, dict]:
    professional_suffix = "1.42" if year >= 2025 else "1.30"
    with tempfile.TemporaryDirectory(prefix="sinopse-eja-") as temp:
        temp_dir = Path(temp)
        eja_title_terms = (
            "matriculas",
            "educacao de jovens e adultos",
            "etapa de ensino",
            "dependencia administrativa",
        )
        eja_workbook = _extract_workbook(
            eja_source,
            (),
            temp_dir,
            title_terms=eja_title_terms,
        )
        professional_workbook = _extract_workbook(
            professional_source, (professional_suffix,), temp_dir
        )
        eja_specs = {
            "mat_eja_total": {
                "required": ("numero de matriculas", "total"),
                "forbidden": ("ensino fundamental", "ensino medio"),
            },
            "mat_eja_fundamental_total": {
                "required": ("ensino fundamental", "total"),
                "forbidden": ("rede publica", "privada"),
            },
            "mat_eja_medio_total": {
                "required": ("ensino medio", "total"),
                "forbidden": ("rede publica", "privada"),
            },
            "mat_eja_fundamental_dep": {
                "dependency_category": ("ensino fundamental",)
            },
            "mat_eja_medio_dep": {
                "dependency_category": ("ensino medio",)
            },
        }
        tecnico_category = (
            ("curso tecnico", "eja", "nivel medio")
            if year >= 2025
            else ("curso tecnico", "integrada a eja")
        )
        professional_specs = {
            "mat_eja_curso_tecnico_integrada": {
                "required": tecnico_category + ("total",),
                "forbidden": ("rede publica", "privada"),
            },
            "mat_eja_fic_integrado_fundamental": {
                "required": ("fic", "eja", "fundamental", "total"),
                "forbidden": ("rede publica", "privada"),
            },
            "mat_eja_fic_integrado_medio": {
                "required": ("fic", "eja", "medio", "total"),
                "forbidden": ("rede publica", "privada"),
            },
            "mat_eja_curso_tecnico_integrada_dep": {
                "dependency_category": tecnico_category
            },
            "mat_eja_fic_integrado_fundamental_dep": {"dependency_category": ("fic", "eja", "fundamental")},
            "mat_eja_fic_integrado_medio_dep": {"dependency_category": ("fic", "eja", "medio")},
        }
        eja, eja_meta = _parse_sheet(
            eja_workbook,
            None,
            year,
            uf,
            eja_specs,
            title_terms=eja_title_terms,
        )
        professional, professional_meta = _parse_sheet(
            professional_workbook, professional_suffix, year, uf, professional_specs
        )
    frame = eja.merge(professional, on=["ano", "id_municipio"], validate="one_to_one")
    frame = frame.rename(
        columns={
            column: column.replace("_dep_", "_")
            for column in frame.columns
            if "_dep_" in column
        }
    )
    denominator = frame["mat_eja_fundamental_total"] + frame["mat_eja_medio_total"]
    if not frame["mat_eja_total"].eq(denominator).all():
        raise ValueError(
            "O Total da EJA diverge de EJA fundamental + EJA médio."
        )
    frame["mat_eja_integrada_educacao_profissional"] = frame[
        [
            "mat_eja_curso_tecnico_integrada",
            "mat_eja_fic_integrado_fundamental",
            "mat_eja_fic_integrado_medio",
        ]
    ].sum(axis=1)
    frame["percentual_eja_integrada_educacao_profissional"] = frame[
        "mat_eja_integrada_educacao_profissional"
    ].div(denominator.where(denominator > 0)).mul(100)
    calculate_eja_integrada_series(frame)
    validate_dependency_totals(frame)
    summary = {
        "year": year,
        "uf": uf,
        "municipalities": len(frame),
        "eja_table": eja_meta["sheet"],
        "professional_table": professional_suffix,
        "eja_source": {"path": str(eja_source), "sha256": sha256(eja_source)},
        "professional_source": {"path": str(professional_source), "sha256": sha256(professional_source)},
        "semantic_resolution": {"eja": eja_meta, "professional": professional_meta},
    }
    return frame, summary


def apply_year(frame: pd.DataFrame, year: int) -> None:
    engine = get_local_postgres_engine()
    extra_columns = [
        "mat_eja_fundamental_total",
        "mat_eja_medio_total",
        *[f"mat_eja_fundamental_{dep}" for dep in DEPENDENCIES],
        *[f"mat_eja_medio_{dep}" for dep in DEPENDENCIES],
    ]
    with engine.begin() as connection:
        for column in extra_columns:
            connection.execute(
                text(f'ALTER TABLE eja_integrada_educacao_profissional ADD COLUMN IF NOT EXISTS "{column}" bigint')
            )
        frame.to_sql("_eja_integrada_sync", connection, if_exists="replace", index=False, method="multi")
        assignments = ", ".join(f'"{column}" = source."{column}"' for column in frame.columns if column not in {"ano", "id_municipio"})
        connection.execute(text(f'''UPDATE eja_integrada_educacao_profissional AS target
            SET {assignments}
            FROM _eja_integrada_sync AS source
            WHERE target.ano = source.ano AND target.id_municipio::text = source.id_municipio::text'''))
        connection.execute(text("DROP TABLE _eja_integrada_sync"))
        updated = connection.execute(
            text("""SELECT count(*)
                    FROM eja_integrada_educacao_profissional
                    WHERE ano = :year
                      AND mat_eja_fundamental_total IS NOT NULL
                      AND mat_eja_medio_total IS NOT NULL"""),
            {"year": year},
        ).scalar_one()
        if updated != EXPECTED_RS_MUNICIPALITIES:
            raise ValueError(f"Escrita incompleta: {updated} municípios atualizados.")


def compare_with_stored_percent(frame: pd.DataFrame, year: int) -> dict:
    engine = get_local_postgres_engine()
    with engine.connect() as connection:
        stored = pd.read_sql_query(
            text("""SELECT id_municipio::text AS id_municipio,
                           percentual_eja_integrada_educacao_profissional AS percentual_anterior
                    FROM eja_integrada_educacao_profissional
                    WHERE ano = :year"""),
            connection,
            params={"year": year},
        )
    compared = frame[["id_municipio", "percentual_eja_integrada_educacao_profissional"]].merge(
        stored, on="id_municipio", how="outer", validate="one_to_one"
    )
    before = pd.to_numeric(compared["percentual_anterior"], errors="coerce")
    after = pd.to_numeric(
        compared["percentual_eja_integrada_educacao_profissional"], errors="coerce"
    )
    changed = ~((before.isna() & after.isna()) | ((before - after).abs() <= 1e-9))
    return {
        "changed_municipality_years": int(changed.sum()),
        "changes": [
            {
                "id_municipio": str(item.id_municipio),
                "before": None if pd.isna(item.percentual_anterior) else float(item.percentual_anterior),
                "after": None
                if pd.isna(item.percentual_eja_integrada_educacao_profissional)
                else float(item.percentual_eja_integrada_educacao_profissional),
            }
            for item in compared.loc[changed].itertuples(index=False)
        ],
    }


def main() -> None:
    parser = argparse.ArgumentParser(description=__doc__)
    parser.add_argument("--year", type=int, required=True)
    parser.add_argument("--eja-input", type=Path, required=True)
    parser.add_argument("--professional-input", type=Path, required=True)
    parser.add_argument("--uf", default="Rio Grande do Sul")
    parser.add_argument("--apply", action="store_true")
    parser.add_argument("--summary-out", type=Path)
    args = parser.parse_args()
    frame, summary = parse_sources(
        args.eja_input, args.professional_input, args.year, args.uf
    )
    summary["database_write"] = "validated_only"
    if args.apply:
        summary["percent_comparison_before_apply"] = compare_with_stored_percent(
            frame, args.year
        )
        apply_year(frame, args.year)
        summary["database_write"] = "applied"
    payload = json.dumps(summary, ensure_ascii=False, indent=2) + "\n"
    print(payload, end="")
    if args.summary_out:
        args.summary_out.parent.mkdir(parents=True, exist_ok=True)
        args.summary_out.write_text(payload, encoding="utf-8")


if __name__ == "__main__":
    main()
