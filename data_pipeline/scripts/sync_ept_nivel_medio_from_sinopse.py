"""Validate and optionally load EPT enrolments from an Inep Sinopse workbook.

The platform only stores municipalities in Rio Grande do Sul.  The collector
uses IBGE municipality codes, never municipality names, and supports the two
known Sinopse layouts:

* 2014-2024: ``Educação Profissional 1.30``;
* 2025 onward: ``Educação Profissional 1.42``.

Run with ``--apply`` only after reviewing the JSON summary printed by the
default validation mode.  The write is a full replacement for the selected
year, inside one database transaction.
"""

from __future__ import annotations

import argparse
import hashlib
import json
import sys
import tempfile
import zipfile
from contextlib import contextmanager
from pathlib import Path
from typing import Iterator

import pandas as pd
from openpyxl import load_workbook
from sqlalchemy import text

BASE_DIR = Path(__file__).resolve().parents[1]
if str(BASE_DIR) not in sys.path:
    sys.path.insert(0, str(BASE_DIR))

from src.data.repository import get_local_postgres_engine  # noqa: E402


RS_MUNICIPALITIES = 497
MODE_PREFIXES = {
    "integrado": "mat_integrado",
    "magisterio": "mat_magisterio",
    "concomitante": "mat_concomitante",
    "subsequente": "mat_subsequente",
    "integrada_eja": "mat_integrada_eja",
}

LAYOUTS = {
    "1.30": {
        "sheet_suffix": "1.30",
        "first_data_row": 11,
        "technical_total_column": None,
        "mode_starts": {
            "integrado": 6,
            "magisterio": 11,
            "concomitante": 16,
            "subsequente": 21,
            "integrada_eja": 26,
        },
    },
    "1.42": {
        "sheet_suffix": "1.42",
        "first_data_row": 14,
        "technical_total_column": 15,
        "mode_starts": {
            "integrado": 16,
            "magisterio": 26,
            "concomitante": 36,
            "subsequente": 46,
            "itinerario_tecnico_exclusivo": 56,
            "integrada_eja": 66,
        },
    },
}


def _layout_for_year(year: int) -> dict:
    return LAYOUTS["1.42" if year >= 2025 else "1.30"]


def _normalise(value: object) -> str:
    return " ".join(str(value or "").strip().split())


def _integer(value: object, *, field: str, row_number: int) -> int:
    if value is None or str(value).strip() == "":
        raise ValueError(f"Valor ausente em {field}, linha {row_number}.")
    try:
        numeric = int(float(value))
    except (TypeError, ValueError) as exc:
        raise ValueError(
            f"Valor invalido em {field}, linha {row_number}: {value!r}."
        ) from exc
    if numeric < 0:
        raise ValueError(f"Valor negativo em {field}, linha {row_number}.")
    return numeric


def _file_sha256(path: Path) -> str:
    digest = hashlib.sha256()
    with path.open("rb") as source:
        for chunk in iter(lambda: source.read(1024 * 1024), b""):
            digest.update(chunk)
    return digest.hexdigest()


def _has_sheet(path: Path, suffix: str) -> bool:
    workbook = load_workbook(path, read_only=True, data_only=True)
    try:
        return any(name.endswith(suffix) for name in workbook.sheetnames)
    finally:
        workbook.close()


@contextmanager
def _workbook_path(source_path: Path, sheet_suffix: str) -> Iterator[Path]:
    if source_path.suffix.lower() == ".xlsx":
        if not _has_sheet(source_path, sheet_suffix):
            raise ValueError(f"A planilha {sheet_suffix} nao foi encontrada em {source_path}.")
        yield source_path
        return

    if source_path.suffix.lower() != ".zip":
        raise ValueError("Informe um arquivo .xlsx ou o .zip oficial da Sinopse.")

    with tempfile.TemporaryDirectory(prefix="sinopse-ept-") as temp_dir:
        with zipfile.ZipFile(source_path) as archive:
            candidates = [
                member
                for member in archive.namelist()
                if member.lower().endswith(".xlsx")
            ]
            for member in candidates:
                candidate_path = Path(temp_dir) / Path(member).name
                with archive.open(member) as origin, candidate_path.open("wb") as target:
                    target.write(origin.read())
                if _has_sheet(candidate_path, sheet_suffix):
                    yield candidate_path
                    return
        raise ValueError(f"Nenhuma planilha contendo {sheet_suffix} foi encontrada no zip.")


def _mode_values(row: tuple, start: int, layout_key: str, row_number: int) -> dict[str, int]:
    def value(column: int, name: str) -> int:
        return _integer(row[column - 1], field=name, row_number=row_number)

    total = value(start, "total")
    if layout_key == "1.30":
        federal = value(start + 1, "federal")
        estadual = value(start + 2, "estadual")
        municipal = value(start + 3, "municipal")
        privada = value(start + 4, "privada")
        publica = federal + estadual + municipal
    else:
        publica = value(start + 1, "publica")
        federal = value(start + 2, "federal")
        estadual = value(start + 3, "estadual")
        municipal = value(start + 4, "municipal")
        privada = value(start + 5, "privada")
        if publica != federal + estadual + municipal:
            raise ValueError(
                f"Rede publica inconsistente na linha {row_number}: "
                f"{publica} != {federal} + {estadual} + {municipal}."
            )
    if total != publica + privada:
        raise ValueError(
            f"Total inconsistente na linha {row_number}: {total} != {publica} + {privada}."
        )
    return {
        "total": total,
        "publica": publica,
        "federal": federal,
        "estadual": estadual,
        "municipal": municipal,
        "privada": privada,
    }


def _record_from_row(
    row: tuple,
    row_number: int,
    layout_key: str,
    year: int,
    *,
    include_municipality_id: bool = True,
) -> dict[str, int | str]:
    layout = LAYOUTS[layout_key]
    values = {
        name: _mode_values(row, start, layout_key, row_number)
        for name, start in layout["mode_starts"].items()
    }
    record: dict[str, int | str] = {"ano": year}
    if include_municipality_id:
        record["id_municipio"] = str(
            _integer(row[3], field="codigo_ibge", row_number=row_number)
        )

    for mode_name, prefix in MODE_PREFIXES.items():
        for dependency, numeric_value in values[mode_name].items():
            record[f"{prefix}_{dependency}"] = numeric_value

    for dependency in ("publica", "federal", "estadual", "municipal", "privada"):
        record[f"mat_ept_nivel_medio_{dependency}"] = sum(
            mode[dependency] for mode in values.values()
        )

    official_total_column = layout["technical_total_column"]
    record["mat_ept_nivel_medio_total"] = (
        _integer(row[official_total_column - 1], field="ept_tecnica_total", row_number=row_number)
        if official_total_column
        else sum(mode["total"] for mode in values.values())
    )
    return record


def parse_workbook(source_path: Path, year: int, uf: str) -> tuple[pd.DataFrame, dict]:
    layout_key = "1.42" if year >= 2025 else "1.30"
    layout = LAYOUTS[layout_key]
    target_uf = _normalise(uf)

    with _workbook_path(source_path, layout["sheet_suffix"]) as workbook_path:
        workbook = load_workbook(workbook_path, read_only=True, data_only=True)
        try:
            sheet_name = next(
                name for name in workbook.sheetnames if name.endswith(layout["sheet_suffix"])
            )
            sheet = workbook[sheet_name]
            records: list[dict[str, int | str]] = []
            state_record: dict[str, int | str] | None = None
            for row_number, row in enumerate(
                sheet.iter_rows(min_row=layout["first_data_row"], values_only=True),
                layout["first_data_row"],
            ):
                if _normalise(row[1]) != target_uf:
                    continue
                if _normalise(row[2]):
                    records.append(_record_from_row(row, row_number, layout_key, year))
                elif row[3] is None or str(row[3]).strip() == "":
                    state_record = _record_from_row(
                        row,
                        row_number,
                        layout_key,
                        year,
                        include_municipality_id=False,
                    )
        finally:
            workbook.close()

    frame = pd.DataFrame(records)
    if frame.empty:
        raise ValueError(f"Nenhum municipio encontrado para a UF {uf!r}.")
    if frame["id_municipio"].duplicated().any():
        duplicates = frame.loc[frame["id_municipio"].duplicated(), "id_municipio"].tolist()
        raise ValueError(f"Codigos IBGE duplicados: {duplicates[:10]}.")
    if len(frame) != RS_MUNICIPALITIES and target_uf == "Rio Grande do Sul":
        raise ValueError(
            f"Cobertura invalida: esperados {RS_MUNICIPALITIES} municipios, encontrados {len(frame)}."
        )
    if state_record is None:
        raise ValueError("Linha de total da UF nao encontrada para validacao.")

    checks = ("mat_ept_nivel_medio_total", "mat_integrado_total", "mat_concomitante_total")
    for field in checks:
        municipality_total = int(frame[field].sum())
        if municipality_total != state_record[field]:
            raise ValueError(
                f"Total da UF divergente em {field}: municipios={municipality_total}, "
                f"Sinopse={state_record[field]}."
            )

    summary = {
        "year": year,
        "layout": layout_key,
        "sheet": sheet_name,
        "uf": uf,
        "municipalities": len(frame),
        "file": str(source_path),
        "sha256": _file_sha256(source_path),
        "totals": {field: int(frame[field].sum()) for field in checks},
    }
    return frame, summary


def replace_year(frame: pd.DataFrame, year: int) -> None:
    engine = get_local_postgres_engine()
    with engine.begin() as connection:
        connection.execute(text("DELETE FROM ept_nivel_medio WHERE ano = :year"), {"year": year})
        frame.to_sql("ept_nivel_medio", connection, if_exists="append", index=False, method="multi")


def main() -> None:
    parser = argparse.ArgumentParser(description=__doc__)
    parser.add_argument("--year", type=int, required=True, help="Ano de referência da Sinopse.")
    parser.add_argument("--input", type=Path, required=True, help="Arquivo .xlsx ou .zip oficial.")
    parser.add_argument("--uf", default="Rio Grande do Sul", help="UF a importar (padrão: Rio Grande do Sul).")
    parser.add_argument("--apply", action="store_true", help="Substitui o ano na tabela ept_nivel_medio.")
    parser.add_argument("--summary-out", type=Path, help="Grava o resumo auditável em JSON.")
    args = parser.parse_args()

    frame, summary = parse_workbook(args.input, args.year, args.uf)
    if args.apply:
        replace_year(frame, args.year)
        summary["database_write"] = "applied"
    else:
        summary["database_write"] = "validated_only"

    payload = json.dumps(summary, ensure_ascii=False, indent=2)
    print(payload)
    if args.summary_out:
        args.summary_out.write_text(payload + "\n", encoding="utf-8")


if __name__ == "__main__":
    main()
